import requests
from allure_commons.types import AttachmentType
from playwright.sync_api import sync_playwright, expect
import openpyxl
import allure




class OpenApplication:



    def read_data_excel(self):
        workbook = openpyxl.load_workbook("../src/files/test_data.xlsx")
        sheet = workbook["sheet_data"]
        total_rows = sheet.max_row
        total_columns = sheet.max_column

        print("rows: ", str(total_rows))
        print("columns:", total_columns)

        userd = sheet.cell(row=2, column=1).value
        passd = sheet.cell(row=2, column=2).value
        user = sheet.cell(row=2, column=3).value
        passw = sheet.cell(row=2, column=4).value
        enviroment = sheet.cell(row=2, column=5).value
        browsers = sheet.cell(row=2, column=6).value
        headless_mode = sheet.cell(row=2, column=7).value
        return userd, passd, user, passw, enviroment,browsers,headless_mode

    @allure.step("Aceder ao ambiente de QA")
    def launch_browser_and_open_page(self):
         self.playwright = sync_playwright().start()
         (_, _, _, _, enviroment, browsers, headless_mode) = self.read_data_excel()
         if browsers.lower() == "chrome":
             self.browser = self.playwright.chromium.launch(headless= headless_mode, channel="chrome")
         elif browsers.lower() == "edge":
             self.browser = self.playwright.chromium.launch(headless=headless_mode, channel="msedge")
         elif browsers.lower() == "firefox":
             self.browser = self.playwright.firefox.launch(headless=headless_mode)
         elif browsers.lower() == "webkit":
             self.browser = self.playwright.webkit.launch(headless=headless_mode)
         else:
             raise ValueError("Navegador inválido")

         # self.context = self.browser.new_context(record_video_dir="videos/")
         # self.page = self.context.new_page()
         self.context = self.browser.new_context()
         self.page = self.browser.new_page()
         # Start tracing before creating / navigating a page.
         # self.context.tracing.start(screenshots=True, snapshots=True, sources=True)
         # Open login page
         self.page.goto(enviroment)
         self.context.tracing.stop(path="trace.zip")
         # allure.attach(self.page.screenshot(path="screenshoot/launchbrowser.png"), name="accessApp", attachment_type=AttachmentType.PNG)
         allure.attach(self.page.screenshot(), name="accessApp",
                       attachment_type=AttachmentType.PNG)
         # HTML
         html_content = "<html><body><h1>Teste2</h1></body></html>"
         allure.attach(html_content, name="Relatório HTML", attachment_type=AttachmentType.HTML)

    def validate_home_page_app(self):
         # self.page.pause()
         # self.page.wait_for_timeout(5000000)
         expect(self.page.locator('//*[@id="root"]/div/div/h1')).to_have_text('Serverest Store')
         # title =  self.page.title()
         # assert title == "Front - ServeRest"
         # expect(self.page).to_have_title("Front - ServeRest")
         text = self.page.locator('//*[@id="root"]/div/div/h1').inner_text()
         print(" >>> Home  as not administrator:", text)

    def close_browser(self):
         self.browser.close()
         self.playwright.stop()
         #pass

    def login_ui(self):
        (_, _, user, passw, _, _, _) = self.read_data_excel()
        print((_, _, user, passw, _, _, _))
        self.page.get_by_test_id("email").wait_for(state="visible", timeout=5000)
        self.page.get_by_test_id("email").fill(user)
        self.page.get_by_test_id("senha").fill(passw)
        # self.page.pause()
        # self.page.wait_for_timeout(5000000)
        self.page.get_by_test_id("entrar").click()
        #self.page.get_by_role("heading", name="Bem Vindo Raul Santos").wait_for(state="visible")
        self.page.get_by_role("heading",name="Serverest Store").wait_for(state="visible")


    def validate_client_area_page(self):
        self.page.wait_for_timeout(2000)
        print('>>>>>>>>>>>>>> validate method area client')
        current_url = self.page.url
        print("Current URL:", current_url)
       #text =  self.page.get_by_role("heading", name="Bem Vindo Raul Santos").inner_text()
        text = self.page.get_by_role("heading",name="Serverest Store").inner_text()
        expect(self.page.get_by_role("heading",name="Serverest Store")).to_have_text("Serverest Store")
        # assert text == "Serverest Store"
        print("assert validation pass----" ,text)


    def login_ui_administrator(self):
            (userd, passd, _, _, _, _, _) = self.read_data_excel()
            # self.page.get_by_test_id("email").wait_for(state="visible", timeout= None)
            self.page.get_by_test_id("email").fill(userd)
            self.page.get_by_test_id("senha").fill(passd)
            
            ###################################################
            btn_visible = self.page.get_by_test_id("entrar").is_visible()
            if not btn_visible:
                    self.page.get_by_test_id("entrar").screenshot(path="files/element.png")
                    allure.attach(self.page.get_by_test_id("entrar").screenshot(), name="button_entrar",
                    attachment_type=AttachmentType.PNG)
                    raise AssertionError("Botão 'Entrar' não está visível")
            ###################################################

            ###### click button with javascript
            self.page.get_by_test_id("entrar").evaluate("element => {element.click()}")
            # self.page.get_by_test_id("entrar").click()

            self.page.wait_for_timeout(2000)
            ###################################################
            self.page.screenshot(path="files/page_after_in.png")
            allure.attach(self.page.screenshot(),name="page_after_in",
                          attachment_type=AttachmentType.PNG)
            ###################################################### fail so com assert senao fica laranja como broke
            # assert  self.page.get_by_test_id("listar-usuarios").is_visible()
            #######################################################
            button = self.page.get_by_test_id("listar-usuarios")
            assert button.is_visible(), "Botão 'listar-usuario' não está visível"
            # allure.attach(self.page.screenshot(), name="fasilurescreeshoot",
            #               attachment_type=AttachmentType.PNG)
            button.click()
            # self.page.screenshot()


            # elements = self.page.locator('//*[@id="root"]/div/div/p/table/tbody').count()
            self.page.locator("table tbody tr").first.wait_for()
            rows = self.page.locator("table tbody tr").count()
            print('elementos:', rows)
            for i in range(rows):
                second_td = self.page.locator(f'//*[@id="root"]/div/div/p/table/tbody/tr[{i + 1}]/td[2]').inner_text()
                print('linhas: ',i,'email: ', second_td)

                if second_td == 'raul.santos@bee-eng.pt':
                  expect(self.page.locator(f'//*[@id="root"]/div/div/p/table/tbody/tr[{i + 1}]/td[2]')).to_have_text('raul.santos@bee-eng.pt')
                  print('>>>>>>>>>>>',second_td)
                  self.page.wait_for_timeout(2000)
                  break
                  ######## comentei pois ta tudo bem mas a pagina nao funiona erro da propria pagina o click button editar
                  # button = self.page.locator(f'//*[@id="root"]/div/div/p/table/tbody/tr[{i+1}]//td[5]]/div/button[1]')
                  # button.scroll_into_view_if_needed()
                  # button.dblclick()


    def api_test_req(self):
        r = requests.post('https://serverest.dev/login', auth=('rsraulsnts@gmail.com', 'JiA4ever#yj12'))
        print(r.status_code)

    def test_api(self):
        # self.playwright = sync_playwright().start()
        self.api = self.playwright.request.new_context(base_url="https://serverest.dev")
        # usa o api_context passado do context do Behave
        response = self.api.post("/login",data={"email": "rsraulsnts@gmail.com","password": "JiA4ever#yj12"})

        assert response.status == 200
        print('<<<', response.status)
        assert response.json()["message"] == "Login realizado com sucesso"
        print('<<<<<',response.json()["message"])
        print('\n',response.json()["authorization"])

        token = response.json()["authorization"]
        print(token)
        # email = "rsraulsnts@gmail.com"
        #
        # # Avalia JS dentro do browser, usando o valor do token
        # self.page.evaluate(
        #     "token => localStorage.setItem('Identity.States.token', token)",token)
        # # Setar email
        # self.page.evaluate(
        #      "email => localStorage.setItem('serverest/userEmail', email)",email)
        self.page.goto('https://front.serverest.dev/home')
        self.page.wait_for_timeout(5000)
        self.api.dispose()
        self.playwright.stop()

    def close_api_all(self):
        self.api.dispose()
        self.playwright.stop()




