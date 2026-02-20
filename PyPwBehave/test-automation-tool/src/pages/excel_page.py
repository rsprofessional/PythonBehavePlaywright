import openpyxl
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parents[1]  # ajusta conforme sua estrutura
EXCEL_PATH = BASE_DIR / "files" / "products.xlsx"


class ReadExcelDinamicaly:
    
    def read_dinamicaly_products(self):
        try:
            workbook = openpyxl.load_workbook(EXCEL_PATH)
            sheet = workbook["sheet_products"]

            products = []  # final list to return

            for row in range(2, sheet.max_row + 1):
                active_product = sheet.cell(row=row, column=1).value

                if active_product:  # só ativos
                    line = []
                    for col in range(2, sheet.max_column + 1):
                        value = sheet.cell(row=row, column=col).value
                        line.append(value)

                    products.append(line)  # guarda linha inteira

            return products

        except Exception as e:
            print("ERRO AO LER EXCEL:", e)
        
