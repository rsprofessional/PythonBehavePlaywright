import openpyxl
from allure_commons.types import AttachmentType
from playwright.sync_api import sync_playwright
import allure


class OpenHerokuapp:

    def read_data_excel_auth_basic(self):
        workbook = openpyxl.load_workbook("../src/files/test_data_herokuap.xlsx")
        sheet = workbook["sheet_data"]
        total_rows = sheet.max_row
        total_columns = sheet.max_column

        #print("rows: ", str(total_rows))
        print("columns:", total_columns)

        username = sheet.cell(row=2, column=1).value
        password = sheet.cell(row=2, column=2).value
        enviroments = sheet.cell(row=2, column=3).value

        return username, password, enviroments

    def auth_basic_poup(self):
        (username, password, _) = self.read_data_excel_auth_basic()

        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(headless=False)
        self.context = self.browser.new_context(http_credentials={"username": username, "password":password})
        self.page =self.context.new_page()
        self.page.goto("https://the-internet.herokuapp.com/digest_auth")
        self.page.wait_for_url("https://the-internet.herokuapp.com/digest_auth")
        self.page.wait_for_timeout(2000)


    def close_browser_auth(self):
         self.browser.close()
         self.playwright.stop()